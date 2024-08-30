import asyncio
import os
import random
from typing import Tuple, Any
import openpyxl

from web3.contract import Contract

from client import Client

rpc = 'https://1rpc.io/linea'
contract_address = '0xDb75Db974B1F2bD3b5916d503036208064D18295'


async def has_name(contract, address) -> bool:
    result = await asyncio.wait_for(contract.functions.redeemed(address).call(), timeout=30)
    return result

async def tmp(address, lines) -> tuple[Any, bool]:
    for i in range(10):
        try:
            random_proxy = random.choice(lines).strip()
            web3 = await Client.configure_web3(rpc, random_proxy)
            break
        except Exception as e:
            if i == 9:
                raise Exception('proxy doesn\'t work')
    abi = [
        {
            "inputs": [
                {
                    "internalType": "address",
                    "name": "_address",
                    "type": "address"
                }
            ],
            "name": "redeemed",
            "outputs": [
                {
                    "internalType": "bool",
                    "name": "",
                    "type": "bool"
                }
            ],
            "stateMutability": "view",
            "type": "function"
        }
    ]
    contract: Contract = web3.eth.contract(abi=abi, address=contract_address)
    res = await asyncio.wait_for(has_name(contract, address), timeout=30)
    return address, res

async def run_tasks_in_batches(tasks, batch_size):
    res = []
    for i in range(0, len(tasks), batch_size):
        batch = tasks[i:i + batch_size]

        results = await asyncio.gather(*batch, return_exceptions=True)
        print(results)
        res += results
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Address"
    sheet["B1"] = "Is named"

    for row_index, result in enumerate(results, start=2):
        if isinstance(result, Exception):
            sheet.cell(row=row_index, column=1).value = "Error"
            sheet.cell(row=row_index, column=3).value = str(result)
        else:
            address, is_named = result
            sheet.cell(row=row_index, column=1).value = address
            sheet.cell(row=row_index, column=2).value = 'YES' if is_named else 'NO'
    workbook.save("people.xlsx")

async def main():
    module_dir = os.path.dirname(__file__)
    filename = module_dir + '/public_keys.txt'
    with open(filename, 'r', encoding='utf-8') as file:
        public_keys = file.readlines()
        public_keys = [key.strip() for key in public_keys]
    module_dir = os.path.dirname(__file__)
    filename = module_dir + '/proxy.txt'
    with open(filename, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    tasks = [tmp(key, lines) for key in public_keys]
    await run_tasks_in_batches(tasks, 100)

if __name__ == "__main__":
    asyncio.run(main())
